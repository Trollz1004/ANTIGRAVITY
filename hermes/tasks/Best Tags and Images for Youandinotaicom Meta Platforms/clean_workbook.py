from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

src = Path('/home/ubuntu/upload/High-Traffic_Social_Communities_and_Dating_App_Mar-Genspark_AI_Sheets-20260403_0344(1).xlsx')
out = Path('/home/ubuntu/youandinotai_clean_review.xlsx')

# Hand-tuned read settings for messy sheets
sheet_configs = {
    'Reddit Communities': {'header': 0},
    'Discord Servers': {'header': 0},
    'Social Media Hashtags': {'header': 0},
    'Launch Directories': {'header': 0},
    '🚀 Submission Tracker': {'header': 3},
    '📅 Content Calendar': {'header': 4},
    '🤖 Prompt Generator': {'header': None},
    'Facebook Groups': {'header': 0},
}

xl = pd.ExcelFile(src)

with pd.ExcelWriter(out, engine='openpyxl') as writer:
    summary_rows = []
    for sheet in xl.sheet_names:
        cfg = sheet_configs.get(sheet, {'header': 0})
        df = pd.read_excel(src, sheet_name=sheet, header=cfg['header'])
        df = df.dropna(how='all')
        if cfg['header'] is None:
            # Preserve prompt sheets as a simple two-column readable format
            df = df.reset_index(drop=True)
            df.columns = [f'column_{i+1}' for i in range(len(df.columns))]
        else:
            df.columns = [str(c).strip() if str(c).strip() != 'nan' else f'column_{i+1}' for i, c in enumerate(df.columns)]
        # Trim totally empty columns after naming
        keep_cols = [c for c in df.columns if not df[c].isna().all()]
        if keep_cols:
            df = df[keep_cols]
        clean_name = sheet[:31]
        df.to_excel(writer, sheet_name=clean_name, index=False)
        summary_rows.append({
            'sheet_name': sheet,
            'clean_sheet_name': clean_name,
            'rows': len(df),
            'columns': len(df.columns),
            'notes': 'Header normalized for readability' if cfg['header'] is not None else 'Kept as prompt/reference sheet'
        })

    pd.DataFrame(summary_rows).to_excel(writer, sheet_name='README', index=False)

wb = load_workbook(out)
header_fill = PatternFill('solid', fgColor='111111')
header_font = Font(color='FFFFFF', bold=True)
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:200]:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
wb.save(out)
print(out)
