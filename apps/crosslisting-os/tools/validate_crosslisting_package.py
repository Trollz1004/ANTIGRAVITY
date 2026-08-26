from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/crosslisting_package/Crosslisting_Review_Template.xlsx')
wb = load_workbook(path, data_only=False)
expected = ['START HERE', 'Listings', 'Reference Lists', 'eBay Upload Review']
print('sheets:', wb.sheetnames)
assert wb.sheetnames == expected, 'Unexpected workbook sheet structure'

listings = wb['Listings']
print('listings_rows_including_header:', listings.max_row)
print('listings_columns:', listings.max_column)
assert listings.max_row == 186, 'Expected one header plus 185 records'
assert listings.max_column == 32, 'Expected 32 listing columns'
assert listings['D2'].value == '012236043904', 'First UPC should preserve leading zeroes'
assert listings['D186'].value == '078693705157', 'Last UPC should use canonical 12-digit UPC'
assert listings['E2'].value in ('', None), 'Placeholder titles must not be copied into listing titles'
assert listings['S2'].value in ('', None), 'Placeholder image URLs must not be copied into listing images'
assert isinstance(listings['AA2'].value, str) and listings['AA2'].value.startswith('=IF('), 'Ready-check formula missing'
assert listings['AD2'].value == 'Not run', 'Extraction status initial state is incorrect'
assert listings['AE2'].value.startswith('DVD Title for UPC '), 'Legacy title reference missing'
assert listings['AF2'].value.startswith('https://example.com/images/'), 'Legacy URL reference missing'
assert listings['AF186'].value in ('', None), 'Known missing image URL should be preserved as blank reference'

preview = wb['eBay Upload Review']
assert preview['A6'].value == '=IF(Listings!$A2="Approved",Listings!C2,"")', 'Preview mapping formula mismatch'

readme = Path('/home/ubuntu/crosslisting_package/README.md').read_text(encoding='utf-8')
script = Path('/home/ubuntu/crosslisting_package/Code.gs').read_text(encoding='utf-8')
assert 'Cross-listing metadata helper' in script
assert 'function onOpen()' in script
assert 'Inventory API' in readme
assert '[1]:' in readme
print('validation: PASS')
