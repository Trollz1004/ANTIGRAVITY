from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SOURCE = Path('/home/ubuntu/upload/crosslistebay')
OUT_DIR = Path('/home/ubuntu/crosslisting_package')
OUT_DIR.mkdir(exist_ok=True)
OUTPUT = OUT_DIR / 'Crosslisting_Review_Template.xlsx'

with SOURCE.open('r', encoding='utf-8', newline='') as f:
    raw_rows = list(csv.DictReader(f))

# Preserve UPC as text and restore leading zeros from the embedded legacy title where available.
def canonical_upc(row):
    title = (row.get('Title') or '').strip()
    prefix = 'DVD Title for UPC '
    if title.startswith(prefix):
        candidate = title[len(prefix):].strip()
        if candidate.isdigit() and 11 <= len(candidate) <= 14:
            return candidate.zfill(12)
    return (row.get('UPC') or '').strip().zfill(12)

records = [{'upc': canonical_upc(row), 'legacy_title': row.get('Title', ''), 'legacy_image': row.get('ImageURL', '')} for row in raw_rows]

wb = Workbook()
ws = wb.active
ws.title = 'START HERE'
listings = wb.create_sheet('Listings')
reference = wb.create_sheet('Reference Lists')
preview = wb.create_sheet('eBay Upload Review')

# Palette
navy = '17365D'
blue = '1F4E78'
light_blue = 'D9EAF7'
light_green = 'E2F0D9'
light_yellow = 'FFF2CC'
light_red = 'FCE4D6'
grey = 'E7E6E6'
white = 'FFFFFF'
dark = '1F1F1F'
thin_grey = Side(style='thin', color='B7B7B7')
border = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)

# START HERE
ws.sheet_view.showGridLines = False
ws.merge_cells('A1:H1')
ws['A1'] = 'Cross-listing Review Template — eBay Preparation'
ws['A1'].font = Font(size=16, bold=True, color=white)
ws['A1'].fill = PatternFill('solid', fgColor=navy)
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 28

intro = [
    ('Purpose', 'Use this workbook as a controlled review queue. Add a source product URL, run the supplied Google Apps Script, verify every extracted field, then transfer approved rows into the current Seller Hub listing template or a separately authenticated API workflow.'),
    ('Important limitation', 'This file prepares listings but does not publish them. Do not use the placeholder titles or example.com image URLs from the original CSV as listing content.'),
    ('Step 1', 'Upload this .xlsx file to Google Drive and open it with Google Sheets. Keep UPC and SKU columns formatted as plain text.'),
    ('Step 2', 'Open Extensions → Apps Script, replace the default file with the supplied Code.gs, and save. Then reload the sheet to enable the Cross-listing menu.'),
    ('Step 3', 'In Listings, add a permitted source URL for each item. Use Cross-listing → Enrich selected rows. The script requests standardized metadata and writes only into the matching row.'),
    ('Step 4', 'Confirm condition, item specifics, category, accurate price, inventory quantity, images, and every seller policy. Change Review Status to Approved only after the row passes review.'),
    ('Step 5', 'Use eBay Upload Review as a staging export. Before any upload, download the current category-specific template from Seller Hub and map these reviewed values into it. Required columns vary by marketplace and category.'),
]
for row_num, (label, text) in enumerate(intro, start=3):
    ws[f'A{row_num}'] = label
    ws[f'A{row_num}'].font = Font(bold=True, color=white)
    ws[f'A{row_num}'].fill = PatternFill('solid', fgColor=blue)
    ws[f'B{row_num}'] = text
    ws[f'B{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'B{row_num}'].fill = PatternFill('solid', fgColor=light_blue if row_num not in (4,) else light_red)
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=8)
    for col in range(1, 9):
        ws.cell(row_num, col).border = border
    ws.row_dimensions[row_num].height = 46

ws['A12'] = 'Workbook tabs'
ws['A12'].font = Font(bold=True, color=white)
ws['A12'].fill = PatternFill('solid', fgColor=navy)
for cell, text in [('A13', 'Listings'), ('B13', 'Working records. Complete only verified data and use its Ready Check field.'),
                   ('A14', 'Reference Lists'), ('B14', 'Editable dropdown choices and default values.'),
                   ('A15', 'eBay Upload Review'), ('B15', 'Approved-row staging table to map into the latest category-specific eBay template.')]:
    ws[cell] = text
for row in range(13, 16):
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill('solid', fgColor=light_green)
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    for col in range(1, 9):
        ws.cell(row, col).border = border

for col, width in {'A': 21, 'B': 28, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 16}.items():
    ws.column_dimensions[col].width = width
ws.freeze_panes = 'A3'

# Reference Lists
reference_headers = ['Review Status', 'Condition', 'Format', 'Region Code', 'Extraction Status', 'Default Value', 'Value']
for col, header in enumerate(reference_headers, 1):
    c = reference.cell(1, col, header)
    c.font = Font(bold=True, color=white)
    c.fill = PatternFill('solid', fgColor=navy)
    c.alignment = Alignment(horizontal='center')
    c.border = border

statuses = ['Needs Source URL', 'Needs Review', 'Research Needed', 'Approved', 'Do Not List', 'Uploaded']
conditions = ['New', 'Like New', 'Very Good', 'Good', 'Acceptable', 'For parts or not working']
formats = ['DVD', 'Blu-ray', '4K UHD', 'VHS', 'Other']
regions = ['Region 0 / All', 'Region 1', 'Region 2', 'Region 4', 'Unknown']
extract_statuses = ['Not run', 'Fetched — review required', 'No structured data found', 'HTTP or access error', 'URL invalid']
defaults = [('Marketplace', 'EBAY_US'), ('Quantity', '1'), ('Currency', 'USD'), ('Condition', 'Good'), ('Format', 'DVD')]
max_len = max(len(statuses), len(conditions), len(formats), len(regions), len(extract_statuses), len(defaults))
for i in range(max_len):
    values = [
        statuses[i] if i < len(statuses) else '',
        conditions[i] if i < len(conditions) else '',
        formats[i] if i < len(formats) else '',
        regions[i] if i < len(regions) else '',
        extract_statuses[i] if i < len(extract_statuses) else '',
        defaults[i][0] if i < len(defaults) else '',
        defaults[i][1] if i < len(defaults) else '',
    ]
    for col, value in enumerate(values, 1):
        cell = reference.cell(i + 2, col, value)
        cell.border = border
for col in range(1, 8):
    reference.column_dimensions[get_column_letter(col)].width = 26
reference.freeze_panes = 'A2'

# Listings
headers = [
    'Review Status', 'Source URL', 'SKU', 'UPC', 'Listing Title', 'Condition', 'Condition Description',
    'Category ID', 'Category Name', 'Brand / Studio', 'Format', 'Region Code', 'Release Year',
    'Director / Key Cast', 'MPN', 'Price (USD)', 'Quantity', 'Description', 'Image URL 1', 'Image URL 2',
    'Image URL 3', 'Image URL 4', 'Fulfillment Policy ID', 'Return Policy ID', 'Payment Policy ID',
    'Merchant Location Key', 'Ready Check', 'Review Notes', 'Last Extracted', 'Extraction Status',
    'Legacy Placeholder Title (reference only)', 'Legacy Placeholder Image URL (reference only)'
]
for col, header in enumerate(headers, 1):
    cell = listings.cell(1, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=navy)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
listings.row_dimensions[1].height = 42

for idx, record in enumerate(records, start=2):
    sku = f'DVD-{record["upc"]}'
    row = [
        'Needs Source URL', '', sku, record['upc'], '', 'Good', '', '', '', '', 'DVD', 'Unknown', '', '', '', '', 1, '', '', '', '', '', '', '', '', '', '', '', '', 'Not run', record['legacy_title'], record['legacy_image']
    ]
    for col, value in enumerate(row, 1):
        cell = listings.cell(idx, col, value)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 7, 18, 28, 31, 32]))
        if col in [4, 3]:
            cell.number_format = '@'
        if col in [16]:
            cell.number_format = '$0.00'
    # Require basic verified fields and seller policies before marking ready.
    listings.cell(idx, 27, f'=IF(AND(A{idx}="Approved",B{idx}<>"",E{idx}<>"",F{idx}<>"",H{idx}<>"",P{idx}>0,Q{idx}>0,R{idx}<>"",S{idx}<>"",W{idx}<>"",X{idx}<>"",Y{idx}<>"",Z{idx}<>""),"READY","NOT READY")')
    listings.cell(idx, 27).font = Font(bold=True)

listings.auto_filter.ref = f'A1:AF{len(records)+1}'
listings.freeze_panes = 'A2'
widths = {
    1: 18, 2: 38, 3: 21, 4: 16, 5: 42, 6: 16, 7: 26, 8: 14, 9: 24, 10: 20, 11: 14, 12: 16,
    13: 14, 14: 26, 15: 16, 16: 14, 17: 10, 18: 48, 19: 38, 20: 38, 21: 38, 22: 38, 23: 22,
    24: 20, 25: 22, 26: 22, 27: 14, 28: 30, 29: 20, 30: 26, 31: 38, 32: 44
}
for col, width in widths.items():
    listings.column_dimensions[get_column_letter(col)].width = width

# Validation lists.
status_dv = DataValidation(type='list', formula1="'Reference Lists'!$A$2:$A$7", allow_blank=False)
condition_dv = DataValidation(type='list', formula1="'Reference Lists'!$B$2:$B$7", allow_blank=False)
format_dv = DataValidation(type='list', formula1="'Reference Lists'!$C$2:$C$6", allow_blank=False)
region_dv = DataValidation(type='list', formula1="'Reference Lists'!$D$2:$D$6", allow_blank=False)
extraction_dv = DataValidation(type='list', formula1="'Reference Lists'!$E$2:$E$6", allow_blank=False)
for dv in [status_dv, condition_dv, format_dv, region_dv, extraction_dv]:
    listings.add_data_validation(dv)
end_row = len(records) + 1
status_dv.add(f'A2:A{end_row}')
condition_dv.add(f'F2:F{end_row}')
format_dv.add(f'K2:K{end_row}')
region_dv.add(f'L2:L{end_row}')
extraction_dv.add(f'AD2:AD{end_row}')

red_fill = PatternFill('solid', fgColor='F4CCCC')
green_fill = PatternFill('solid', fgColor='D9EAD3')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')
listings.conditional_formatting.add(f'AA2:AA{end_row}', FormulaRule(formula=['AA2="READY"'], fill=green_fill))
listings.conditional_formatting.add(f'AA2:AA{end_row}', FormulaRule(formula=['AA2="NOT READY"'], fill=red_fill))
listings.conditional_formatting.add(f'A2:A{end_row}', FormulaRule(formula=['A2="Approved"'], fill=green_fill))
listings.conditional_formatting.add(f'A2:A{end_row}', FormulaRule(formula=['A2="Needs Review"'], fill=yellow_fill))

# eBay Upload Review
preview.sheet_view.showGridLines = False
preview.merge_cells('A1:M1')
preview['A1'] = 'Approved Listing Staging — Map Into a Current Seller Hub Template'
preview['A1'].font = Font(size=14, bold=True, color=white)
preview['A1'].fill = PatternFill('solid', fgColor=navy)
preview['A1'].alignment = Alignment(horizontal='left')
preview.row_dimensions[1].height = 26
preview['A3'] = 'Only rows marked Approved in Listings will appear below. Validate the destination template before uploading; eBay requirements vary by category and marketplace.'
preview['A3'].font = Font(italic=True, color=dark)
preview['A3'].fill = PatternFill('solid', fgColor=light_yellow)
preview.merge_cells('A3:M3')
preview_headers = ['SKU', 'UPC', 'Title', 'Condition', 'Category ID', 'Price', 'Quantity', 'Description', 'Image URL 1', 'Fulfillment Policy ID', 'Return Policy ID', 'Payment Policy ID', 'Merchant Location Key']
for col, header in enumerate(preview_headers, 1):
    cell = preview.cell(5, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=blue)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border
for row in range(2, end_row + 1):
    pr = row + 4
    # Use formulas so Google Sheets/Excel users see review output dynamically.
    mapping = {1: 'C', 2: 'D', 3: 'E', 4: 'F', 5: 'H', 6: 'P', 7: 'Q', 8: 'R', 9: 'S', 10: 'W', 11: 'X', 12: 'Y', 13: 'Z'}
    for col, source_col in mapping.items():
        cell = preview.cell(pr, col, f'=IF(Listings!$A{row}="Approved",Listings!{source_col}{row},"")')
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in [3, 8, 9]))
        if col == 6:
            cell.number_format = '$0.00'
for col in range(1, 14):
    preview.column_dimensions[get_column_letter(col)].width = 18
preview.column_dimensions['C'].width = 38
preview.column_dimensions['H'].width = 46
preview.column_dimensions['I'].width = 38
preview.freeze_panes = 'A6'
preview.auto_filter.ref = f'A5:M{end_row+4}'

# Sheet tabs and protection cues.
for sheet in [ws, listings, reference, preview]:
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.sheet_view.zoomScale = 85

wb.save(OUTPUT)
print(f'Created {OUTPUT}')
print(f'Imported {len(records)} source rows')
print(f'Rows without legacy image URL: {sum(1 for r in records if not r["legacy_image"])}')
